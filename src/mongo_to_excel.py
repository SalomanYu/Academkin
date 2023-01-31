import pymongo
import openpyxl


client = pymongo.MongoClient('localhost', 27017)
db = client["academkin"]
vuzCol = db["Vuz"]
specCol = db["Specialization"]


def exportVuzes():
    book = openpyxl.Workbook()
    sheet = book.create_sheet("vuzes")
    rowCount = 2 
    for item in vuzCol.find():
        sheet.cell(rowCount, 1).value = item["vuz_id"]
        sheet.cell(rowCount, 2).value = item["short_name"]
        sheet.cell(rowCount, 3).value = item["full_name"]
        sheet.cell(rowCount, 4).value = item["logo"]
        sheet.cell(rowCount, 5).value = item["city"]
        sheet.cell(rowCount, 6).value = item["locality"]
        rowCount += 1
    print("Done for vuzes")
    book.save("academkin.xlsx")

def exportSpecs():
    book = openpyxl.load_workbook("academkin.xlsx")
    sheet = book.create_sheet("specs")
    rowCount = 2 
    for item in specCol.find():
        sheet.cell(rowCount, 1).value = item["spec_id"]
        sheet.cell(rowCount, 2).value = item["vuz_id"]
        sheet.cell(rowCount, 3).value = item["vuz_name"]
        sheet.cell(rowCount, 4).value = item["name"]
        sheet.cell(rowCount, 5).value = item["form_educations"]
        sheet.cell(rowCount, 6).value = item["duration"]
        sheet.cell(rowCount, 7).value = item["preparation_level"]
        sheet.cell(rowCount, 8).value = item["qualification"]
        rowCount += 1
    print("Done for specs")
    book.save("academkin.xlsx")

if __name__ == "__main__":
    exportVuzes()
    exportSpecs()